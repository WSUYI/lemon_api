from openpyxl import load_workbook
import requests

# 读取
def read_case(excel_file,sheet_name):
    wb = load_workbook(excel_file)
    sh = wb[sheet_name]
    rows =sh.max_row
    # cl=sh.cell(row=1,column=1).value
    # print(cl)
    # url = sh.cell(row=2,column=5).value
    # data = sh.cell(row=2,column=6).value
    # expected = sh.cell(row=2,column=7).value
    # print(url,data,expected)
    list = []
    for i in range(2,rows + 1):
        case_dict = dict(
            case_id = sh.cell(row=i, column=1).value,
            url = sh.cell(row=i, column=5).value,
            data = sh.cell(row=i, column=6).value,
            expected = sh.cell(row=i, column=7).value
        )
        list.append(case_dict)
        # print(case_id,url, data, expected)
    return list
# 写入
def write_case(file_name,sheet_name,row,column,value):
    wb = load_workbook(file_name)
    sh = wb[sheet_name]
    sh.cell(row=row, column=column).value = value
    wb.save(file_name)
# 请求
def jiekou(url,data):
    header = {'X-Lemonban-Media-Type':'lemonban.v2','Content-Type':'application/json'}
    res= requests.post(url=url, json=data, headers=header)
    #print(res_login.json())
    return res.json()